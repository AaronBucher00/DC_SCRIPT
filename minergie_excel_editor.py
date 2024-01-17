import os #for acess on explorer
import pandas as pd #for edit excel
import openpyxl #for edit excel
import warnings #for warning messages in tkinter
import shutil #for copy original excel

import tkinter as tk #for UI
from tkinter import filedialog, StringVar, OptionMenu, Label, Button, Toplevel

import win32com.client #for update values in exported Minergie-Excels

def minergie_excel_editor(glb_exp_data):
    class Minergie_Excel_Editor:
        def __init__(self, parent):
            self.root = parent
            self.root.title("Minergie Excel Editor")

            # Main window components
            self.load_button = tk.Button(root, text="Vorlage Excel von Minergie öffnen", command=self.load_excel)
            self.load_button.pack(pady=10)

            self.file_label = tk.Label(root, text="Keine Datei ausgewählt")
            self.file_label.pack(pady=10)

            self.frames = [
            self.create_project_info_frame(),
            self.create_room_info_frame(),
            self.create_windows_info_frame(),
            self.create_sonnenschutz_frame(),
            self.create_komfort_frame()
            ]

            self.current_frame = 0
            self.show_frame()

            self.status_label = tk.Label(root, text="")
            self.status_label.pack(pady=10)


        def show_frame(self):
            for i, frame in enumerate(self.frames):
                if i == self.current_frame:
                    frame.pack()
                else:
                    frame.pack_forget()


        def next_frame(self):
            self.current_frame += 1
            if self.current_frame < len(self.frames):
                self.show_frame()


        def prev_frame(self):
            self.current_frame -= 1
            if self.current_frame >= 0:
                self.show_frame()

        
        def load_excel(self):
            original_filepath = filedialog.askopenfilename(filetypes=[
                ("Excel Dateien", "*.xlsx;*.xlsm;*.xlsb;*.xltx;*.xltm;*.xls;*.xlt;*.xml;*.xlam;*.xla;*.xlw;*.xlr")])

            if original_filepath:
                self.file_label.config(text=original_filepath)
                self.workbook = openpyxl.load_workbook(original_filepath)
                self.save_button.config(state=tk.NORMAL)
            else:
                self.status_label.config(text="Fehler: Keine Datei ausgewählt", fg="red")

        
        def create_project_info_frame(self): ##### PROJEKT INFORMATIONEN #####
            frame = tk.Frame(self.root)
            frame.pack(pady=10)
            tk.Label(frame, text="PROJEKT INFORMATIONEN").grid(row=0, column=0, columnspan=2)

            project_info_entry = ["Projektname", "Parz. Nr.", "MOP - Nr.", "Gebäudeadresse"]
            project_info_drop = ["Klimastation", "Lage des Projektes", "Lage in Föhngebiet", "Klimadaten", "Wärmeinseleffekt"]

            klimastation = ["Adelboden", "Aigle", "Altdorf", "Basel-Binningen", "Bern Liebefeld", "Buchs-Aarau", "Chur", "Davos", "Disentis", "Engelberg", "Genève-Cointrin", "Glarus", "Grand-St-Bernard", "Güttingen", "Interlaken",
                            "La Chaux-de-Fonds", "La Frétaz", "Locarno-Monti", "Lugano", "Luzern", "Magadino", "Montana", "Neuchâtel", "Payerne", "Piotta", "Pully", "Robbia", "Rünenberg", "Samedan", "San Bernardino", "St. Gallen", "Schaffhausen",
                            "Scuol", "Sion", "Ulrichen", "Vaduz", "Wynau", "Zermatt", "Zürich-Kloten", "Zürich-MeteoSchweiz"]
            lage_projektes = ["Seeufer", "Grosse Ebene", "Ortschaften, freies Feld", "Grossflächige Stadtgebiete"]
            lage_foehngebiet = ["Ja", "Nein"]
            klimadaten = ["2010", "2035", "2060 low", "2060 high"]
            waermeinseleffekt = ["Ja", "Nein"]

            project_info_lists = [klimastation, lage_projektes, lage_foehngebiet, klimadaten, waermeinseleffekt]

            self.list_project_info = []

            # EINGABE
            for i, label in enumerate(project_info_entry):
                tk.Label(frame, text=label).grid(row=i + 1, column=0)
                entry = tk.Entry(frame)
                entry.grid(row=i + 1, column=1)
                self.list_project_info.append(entry)

            # DROP-DOWN
            for j, options in enumerate(project_info_lists):
                selection = StringVar()
                selection.set(options[0])
                tk.Label(frame, text=project_info_drop[j]).grid(row=j + len(project_info_entry) + 1, column=0)
                drop = OptionMenu(frame, selection, *options)
                drop.grid(row=j + len(project_info_entry) + 1, column=1)
                self.list_project_info.append(selection)

            tk.Button(frame, text="Weiter", command=self.next_frame).grid(row=len(project_info_entry) + len(project_info_lists) + 2, columnspan=2, pady=10)
            return frame


        def create_room_info_frame(self):  ##### EINGABEN WÄRMESPEICHERFÄHIGKEIT DER BAUTEILE #####
            frame = tk.Frame(self.root)
            frame.pack(pady=10)
            tk.Label(frame, text="EINGABEN WÄRMESPEICHERFÄHIGKEIT DER BAUTEILE").grid(row=0, column=0, columnspan=2)

            room_info = ["Auswahl Boden", "Auswahl Decke", "Auswahl Innenwand", "Auswahl Aussenwand opak"]

            boden_options = ["Parkett auf Unterlagsboden > 6cm", "Platten/Keramik auf Unterlagsboden", "Teppich auf Doppelboden"]
            decke_options = ["Massivdecke 24cm", "Holzbalkendecke EFH", "Holzbalkendecke MFH","HBV - Brettstapeldecke / Beton", "Vollholzdecke 15cm", "Akustikmassnahmen auf Massivdecke"]
            innenwand_options = ["Beton 20cm verputzt", "GK-Leichtbauwand 3-fach beplankt", "GK-Leichtbauwand 2-fach beplankt", "GK-Leichtbauwand 1-fach beplankt", "Mauerwerk 12-18cm verputzt"]
            aussenwand_options = ["raumseitig Beton 20cm, verputzt", "Holzständerwand", "Mauerwerk 12-18cm verputzt", "Innendämmung u. Vorsatzschale"]

            room_info_lists = [boden_options, decke_options, innenwand_options, aussenwand_options]

            self.list_room_info = []

            # DROP-DOWN
            for i, options in enumerate(room_info_lists):
                selection = StringVar()
                selection.set(options[0])
                tk.Label(frame, text=room_info[i]).grid(row=i + 1, column=0)
                drop = OptionMenu(frame, selection, *options)
                drop.grid(row=i + 1, column=1)
                self.list_room_info.append(selection)

            tk.Button(frame, text="Zurück", command=self.prev_frame).grid(row=len(room_info) + 2, columnspan=2, pady=10)
            tk.Button(frame, text="Weiter", command=self.next_frame).grid(row=len(room_info) + 3, columnspan=2, pady=10)

            return frame


        def create_windows_info_frame(self): ##### FENSTER EIGENSCHAFTEN #####
            frame = tk.Frame(self.root)
            frame.pack(pady=10)
            tk.Label(frame, text="FENSTER EIGENSCHAFTEN").grid(row=0, column=0, columnspan=2)

            windows_info_entry = ["Parz. Nr.", "MOP - Nr.", "Gebäudeadresse"]

            windows_info_drop = ["Rahmenanteil Fenster", "Refkletion Fassade gegenüber", "g-Wert Verglasung", "g-Wert Total", "Horizontwinkel (in Grad)", "Bauliche Verschattung Überhang (in cm)", "Bauliche Verschattung Leibung (in cm)"]

            rahmenanteil_fenster = ["0.25", "0.24", "0.23", "0.22", "0.21", "0.20", "0.19", "0.18", "0.17", "0.16", "0.15", "0.14", "0.13", "0.12", "0.11", "0.10", "0.09", "0.08", "0.07", "0.06", "0.05"]
            reflektion_fassade_gg = ["Ja", "Nein"]
            g_wert_glas = ["0.53", "0.47", "0.42", "0.38", "0.34", "0.30", "0.26", "0.24", "0.22", "0.20"]
            g_wert_total = ["0.15", "0.14", "0.13", "0.12", "0.11", "0.10", "0.09", "0.08", "0.07", "0.06", "0.05"]
            horizontwinkel = ["0", "2", "5", "10", "20", "30", "40", "50", "60", "70", "80", "90"]
            bauliche_verschattung_ueberhang = ["gem. IFC-Modell","0.50", "0.40", "0.35", "0.30", "0.25", "0.20", "0.15", "0.10", "0.05", "0.00"]
            bauliche_verschattung_leibung = ["gem. IFC-Modell","0.50", "0.40", "0.35", "0.30", "0.25", "0.20", "0.15", "0.10", "0.05", "0.00"]

            windows_info_lists = [rahmenanteil_fenster, reflektion_fassade_gg, g_wert_glas, g_wert_total, horizontwinkel, bauliche_verschattung_ueberhang, bauliche_verschattung_leibung]

            self.list_windows_info = []

            # DROP-DOWN 
            for i, options in enumerate(windows_info_lists):
                selection = StringVar()
                selection.set(options[0])
                tk.Label(frame, text=windows_info_drop[i]).grid(row=i + 1, column=0)
                drop = OptionMenu(frame, selection, *options)
                drop.grid(row=i + 1, column=1)
                self.list_windows_info.append(selection)

            tk.Button(frame, text="Zurück", command=self.prev_frame).grid(row=len(windows_info_drop) + 2, columnspan=2, pady=10)
            tk.Button(frame, text="Weiter", command=self.next_frame).grid(row=len(windows_info_drop) + 3, columnspan=2, pady=10)
            
            return frame


        def create_sonnenschutz_frame(self): ##### WINDFESTIGKEIT DER SONNENSCHUTZEINRICHTUNG #####
            frame = tk.Frame(self.root)
            frame.pack(pady=10)
            tk.Label(frame, text="WINDFESTIGKEIT DER SONNENSCHUTZEINRICHTUNG").grid(row=0, column=0, columnspan=2)

            sonnenschutz_info = ["Minergie Modul Sonnenschutz verwendet?", "mindest empfohlene Windwiderstandsklasse umgesetzt?", "externer Nachweis mit niedrigerer Windwiderstandsklasse wird geführt?", "Deklaration des geplanten Sonnenschutzes"]

            minergie_modul = ["Ja", "Nein"]
            mind_widerstandsklasse = ["Ja", "Nein"]
            ect_nachweis_nid_widerstandsklasse = ["Ja", "Nein"]
            deklaration_sonnenschutz = ["Ausstellmarkise", "Senkrechtmarkise", "Horizontalmarkise", "Fallarmmarkise", "Lamellenstoren", "Rollladen", "Fensterladen", "Schiebeladen", "kein Sonnenschutz"]
            
            sonnenschutz_info_lists = [minergie_modul, mind_widerstandsklasse, ect_nachweis_nid_widerstandsklasse, deklaration_sonnenschutz]

            self.list_sonnenschutz = []

            # EINGABE
            tk.Label(frame, text="Einbauhöhe der Sonnenschutzeinrichtung (mind. 2.5m)").grid(row=1, column=0)
            entry = tk.Entry(frame)  # empty space for the user to write in
            entry.grid(row=1, column=1)
            self.list_sonnenschutz.append(entry)

            # DROP-DOWN 
            for i, options in enumerate(sonnenschutz_info_lists):
                selection = StringVar()
                selection.set(options[0])
                tk.Label(frame, text=sonnenschutz_info[i]).grid(row=i + 2, column=0)
                drop = OptionMenu(frame, selection, *options)
                drop.grid(row=i + 2, column=1)
                self.list_sonnenschutz.append(selection)

            tk.Button(frame, text="Zurück", command=self.prev_frame).grid(row=len(sonnenschutz_info) + 3, columnspan=2, pady=10)
            tk.Button(frame, text="Weiter", command=self.next_frame).grid(row=len(sonnenschutz_info) + 4, columnspan=2, pady=10)
           
            return frame


        def create_komfort_frame(self): ##### ABFRAGE ZUM SOMMERLICHEN KOMFORT #####
            frame = tk.Frame(self.root)
            frame.pack(pady=10)
            tk.Label(frame, text="ABFRAGE ZUM SOMMERLICHEN KOMFORT").grid(row=0, column=0, columnspan=2)

            komfort_info = ["Auswahl Nutzungskategorie", "Auswahl Sommerstrategie"]

            nutzungskategorie = ["MFH (SIA 2024:2015)", "EFH (SIA 2024:2015)", "Hotelzimmer (SIA 2024:2015)", "Einzel-, Gruppenbüro (SIA 2024:2015)", "Grossraumbüro (SIA 2024:2015)", "Sitzungszimmer (SIA 2024:2015)", "Schulzimmer (SIA 2024:2015)",
                                "Lehrerzimmer (SIA 2024:2015)", "Bibliothek (SIA 2024:2015)", "Hörsaal (SIA 2024:2015) ", "Schulfachraum (SIA 2024:2015)", "Lebensmittelverkauf (SIA 2024:2015)", "Fachgeschäft (SIA 2024:2015)", "Verkauf Möbel, Bau, Gar. (SIA 2024:2015)", 
                                "Restaurant (SIA 2024:2015)", "Restaurant Selbstbedien. (SIA 2024:2015)", "Mehrzweckhalle (SIA 2024:2015)", "Ausstellungshalle (SIA 2024:2015)", "Bettenzimmer (SIA 2024:2015)", "Stationszimmer (SIA 2024:2015)",
                                "Behandlungsraum (SIA 2024:2015)", "Produktion (grob) (SIA 2024:2015)", "Produktion (fein) (SIA 2024:2015)", "Laborraum (SIA 2024:2015)", "Lagerhalle (SIA 2024:2015)", "Turnhalle (SIA 2024:2015)", "Fitnessraum (SIA 2024:2015)"]
            sommerstrategie = ["Fensterlüftung Tag", "Fensterlüftung Tag&Nacht", "Fensterquerlüftung Tag&Nacht", "Fensterlüftung Tag & mech.Lüftung Nacht", "mech.Lüftung (inkl. Nacht) mit Sommer Bypass", "FBK-FreeCooling&mech.Lüftung mit Sommerbypass (inkl. Nacht)", "FBK-FreeCooling & Fensterlüftung"]
            
            komfort_info_lists = [nutzungskategorie, sommerstrategie]

            self.list_komfort = []
        
            # DROP-DOWN 
            for i, options in enumerate(komfort_info_lists):
                selection = StringVar()
                selection.set(options[0])
                tk.Label(frame, text=komfort_info[i]).grid(row=i + 1, column=0)
                drop = OptionMenu(frame, selection, *options)
                drop.grid(row=i + 1, column=1)
                self.list_komfort.append(selection)
            

            tk.Button(frame, text="Zurück", command=self.prev_frame).grid(row=len(komfort_info_lists) + 1, column=0, pady=10)
            self.save_button = tk.Button(frame, text="Speicherort Excel Files auswählen", command=self.save_values, state=tk.DISABLED)
            self.save_button.grid(row=len(komfort_info_lists) + 2, column=0, pady=10)
            
            return frame 
        

        def save_values(self):
            try:
                save_folder = filedialog.askdirectory()# Let the user choose the destination folder

                if save_folder:
                    
                    room_nr = 0
                    file_nr = 0
                    save_values_done = False

                    while True:    
                        # Create the full path for the saved Excel file
                        file_nr +=1
                        save_filepath = os.path.join(save_folder, "sows_v2023-2_"+ str(file_nr) +".xlsx")

                        self.sheet_name_entry_info = "Projekteingaben"
                        info_sheet = self.workbook[self.sheet_name_entry_info]  # Access the sheet by its name

                        #PROJEKTEINGABEN
                        info_sheet['D7'] = self.list_project_info[0].get()  # Projektname
                        info_sheet['H7'] = self.list_project_info[1].get()  # Parz. Nr.
                        info_sheet['J7'] = self.list_project_info[2].get()  # MOP - NR.
                        info_sheet['D8'] = self.list_project_info[3].get()  # Gebäudeadresse
                        info_sheet['F11'] = self.list_project_info[4].get()  # Klimastation
                        info_sheet['F12'] = self.list_project_info[5].get()  # Lage des Projektes
                        info_sheet['F13'] = self.list_project_info[6].get()  # Lage in Föhngebiet
                        info_sheet['J11'] = self.list_project_info[7].get()  # Klimadaten
                        info_sheet['J13'] = self.list_project_info[8].get()  # Wärmeinseleffekt

                        for p in range(3):
                            self.sheet_name_entry_room = "Nachweisblatt_Raum." + str(p+1)
                            room_sheet = self.workbook[self.sheet_name_entry_room]  # Access the sheet by its name 
                            
                            #EINGABEN RAUM
                            room_sheet['E11'] = glb_exp_data[room_nr][0] + "_" + glb_exp_data[room_nr][1] # Raumbezeichnung
                            room_sheet['L11'] = glb_exp_data[room_nr][2] # relevante NGF Raum
                            
                            #EINGABEN WÄRMESPEICHERFÄHIGKEIT
                            room_sheet['E18'] = self.list_room_info[0].get()  # Auswahl Boden: Parkett auf Unterlagsboden > 6cm
                            room_sheet['H18'] = glb_exp_data[room_nr][2]  # Fläche Boden
                            room_sheet['E19'] = self.list_room_info[1].get()  # Auswahl Decke: Vollholzdecke 15cm
                            room_sheet['H19'] = glb_exp_data[room_nr][2]  # Fläche Decke
                            room_sheet['E20'] = self.list_room_info[2].get()  #A uswahl Innenwand: GK-Leichtbauwand 2-fach beplankt
                            room_sheet['H20'] = "=L11^(1/2)*"+ str(glb_exp_data[room_nr][3]) +"-2.0"  # Fläche Innenwand
                            room_sheet['E21'] = self.list_room_info[3].get()  # Auswahl Aussenwand: Holzständerwand
                            room_sheet['H21'] = "=L11^(1/2)*"+ str(glb_exp_data[room_nr][3]) +"-(E31*E32*E33)-(H31*H32*H33)-(K31*K32*K33)"  # Fläche Aussenwand

                            #EINGABEN FENSTER UND BAULICHE VERSCHATTUNG              
                            for j in range(0,9,4):
                                if j == 0:  # Fenstertyp FE01                  
                                    fenstertyp = "FE01"
                                    spalte_fenstertyp = "F"
                                    spalte = "E"
                                elif j == 4:  # Fenstertyp FE02
                                    fenstertyp = "FE02"
                                    spalte_fenstertyp = "I"
                                    spalte = "H"
                                elif j == 8:  # Fenstertyp FE03
                                    fenstertyp = "FE03"
                                    spalte_fenstertyp = "L"
                                    spalte = "K"
                                else:
                                    print("ERRROR - check window types")
                                

                                if glb_exp_data[room_nr][5+j] == 0:
                                    pass
                                else:
                                    room_sheet[spalte_fenstertyp + '28'] = fenstertyp  # Fenstertyp
                                    room_sheet[spalte + '29'] = glb_exp_data[room_nr][5+j]  # Ausrichtung
                                    if fenstertyp == "FE01":
                                        pass
                                    else:
                                        room_sheet[spalte + '30'] = "Nein"  # Neigung
                                    #room_sheet[spalte_fenstertyp + '30'] = "90" # Grad der Neigung                                                                                               
                                
                                
                                if glb_exp_data[room_nr][5+j] == 0:
                                    #room_sheet[spalte + '31'] = int(0)  # Fensteranzahl
                                    pass
                                else:                      
                                    room_sheet[spalte + '31'] = int(1)  # Fensteranzahl

                                    room_sheet[spalte + '32'] = float(glb_exp_data[room_nr][6+j])  # Fensterbreite
                                    room_sheet[spalte + '33'] = float(glb_exp_data[room_nr][7+j])  # Fensterhöhe
                                    room_sheet[spalte + '34'] = float(self.list_windows_info[0].get())  # Rahmenanteil

                                    room_sheet[spalte + '38'] = "=" + str(spalte) + '33/2.0'  # Abstand Überhang

                                    if self.list_windows_info[5].get() == "gem. IFC-Modell" :
                                        room_sheet[spalte + '39'] = float((glb_exp_data[room_nr][8+j]))  # Länge Überhang
                                    else:
                                        room_sheet[spalte + '39'] = float(self.list_windows_info[5].get())  # Länge Überhang
                                    
                                    room_sheet[spalte + '40'] = "=" + str(spalte) + '32/2.0'  # Abstand Seitenblende rechts
                                    room_sheet[spalte + '42'] = "=" + str(spalte) + '32/2.0'  # Abstand Seitenblende links

                                    if self.list_windows_info[6].get() == "gem. IFC-Modell" :
                                        room_sheet[spalte + '41'] = float((glb_exp_data[room_nr][8+j]))  # Länge Seitenblende rechts
                                        room_sheet[spalte + '43'] = float((glb_exp_data[room_nr][8+j]))  # Länge Seitenblende links
                                    else:
                                        room_sheet[spalte + '41'] = float(self.list_windows_info[6].get())  # Länge Seitenblende rechts
                                        room_sheet[spalte + '43'] = float(self.list_windows_info[6].get())  # Länge Seitenblende links

                                    room_sheet[spalte + '44'] = int(self.list_windows_info[4].get())  # Horizontwinkel
                                    room_sheet[spalte + '45'] = str(self.list_windows_info[1].get())  # Reflexion Fassade gegenüber
                                    room_sheet[spalte + '47'] = float(self.list_windows_info[2].get())  # g-Wert Verglasung

                                    room_sheet[spalte + '56'] = float(self.list_windows_info[3].get())  # g-total
                            

                            #Windfestigkeit der Sonnenschutzeinrichtung
                            room_sheet['E60'] = self.list_sonnenschutz[0].get()  # Einbauhöhe
                            room_sheet['L60'] = self.list_sonnenschutz[1].get()  # Minergie Modul Sonnenschutz verwendet
                            room_sheet['L66'] = self.list_sonnenschutz[2].get()  # mind. empfohlene Widerstandsklasse umgesetzt
                            room_sheet['L67'] = self.list_sonnenschutz[3].get()  # Nachweis mit nidrigerer Widerstandsklasse geführt
                            room_sheet['K69'] = self.list_sonnenschutz[4].get()  # Deklaration geplanter Sonnenschutz

                            #Abfrage zum sommerlichen Komfort
                            room_sheet['D94'] = self.list_komfort[0].get()  # Nutzungskategorie
                            room_sheet['D96'] = self.list_komfort[1].get()  # Sommerstrategie

                            print("INFO - Room Nr. "+ str(room_nr) + "/"+str(len(glb_exp_data)-1) + " saved") 

                            if (room_nr) >= len(glb_exp_data)-1:
                                save_values_done = True
                                break 
                            else:
                                room_nr += 1   

                        self.workbook.save(save_filepath)

                        if (room_nr-1) >= len(glb_exp_data)-1 or save_values_done == True:
                            save_status = tk.Label(root, text="Excel Files erfolgreich gespeichert", fg="green")
                            save_status.pack(pady=10)
                            break
                                                    
                else:
                    save_status = tk.Label(root, text="Fehler: Kein Speicherort ausgewählt", fg="red")
                    save_status.pack(pady=10)
                
            except Exception as e:
                self.status_label.config(text=f"Fehler: {str(e)}", fg="red")

            def recalculate_excel_file(file_path):
                excel_app = win32com.client.DispatchEx("Excel.Application")  # Use DispatchEx for better control
                excel_app.Visible = False  # Set visibility to False

                workbook = None  # Initialize workbook with None

                try:
                    # Open the Excel file
                    workbook = excel_app.Workbooks.Open(file_path)
                                
                    # Recalculate all formulas
                    workbook.RefreshAll()
                                
                    # Save the changes
                    workbook.Save()
                                
                except Exception as e:
                    print(f"Error: {str(e)}")

                finally:
                    if workbook is not None:
                        # Close Excel
                        workbook.Close(SaveChanges=False)  # SaveChanges=False prevents saving changes
                    excel_app.Quit()
                            

            def total_files_finder(filepath):
                punkt_index = filepath.rfind('.')
                if punkt_index != -1:
                    zeichen_vor_punkt = filepath[punkt_index - 1]
                else:
                    zeichen_vor_punkt = "ERROR - check function total_files_finder"

                return zeichen_vor_punkt
                    

            def generate_temp_save_filepath(filepath, current_file_nr):
                punkt_index = filepath.rfind('.')
                if punkt_index != -1 and punkt_index >= len(filepath) - 6:
                    temp_save_filepath = filepath[:punkt_index - 1] + str(current_file_nr) + filepath[punkt_index:]
                else:
                    print("ERROR - check function save_filepath")

                return temp_save_filepath
                        
                        
            def save_overview_info_to_excel(overview_info, output_filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active

                # Write headers to the first row of the sheet
                headers = ['Excel-Pfad','Raumname', 'Einhaltung bauliche Grundanforderungen', 'Einhaltung sommerlichen Komfort', 'Benötigte Maßnahmen', 'Benötigte g-tot-Werte']
                sheet.append(headers)

                # Write data to the sheet
                for room_info in overview_info:
                    sheet.append(room_info)

                # Save the workbook to the specified file path
                workbook.save(output_filepath)
                        

            def save_overview(save_folder,save_filepath):
                                        
                total_files = int(total_files_finder(save_filepath))
                total_rooms = int(len(glb_exp_data)-1)
                current_room = 0

                            
                overview_info = []

                for i in range(1,(total_files+1)):
                    print("INFO - File Nr.", i, "read")

                            
                temp_save_filepath = generate_temp_save_filepath(save_filepath, i)

                try:
                    if temp_save_filepath:
                        self.file_label.config(text=temp_save_filepath)
                        # Recalculate formulas in Excel file
                        recalculate_excel_file(temp_save_filepath)
                        self.workbook = openpyxl.load_workbook(temp_save_filepath, data_only=True)
                        self.save_button.config(state=tk.NORMAL)
                    else:
                        self.status_label.config(text="Error: No file selected", fg="red")

                                    
                    for j in range(1,4):
                        if current_room >= (total_rooms+1):
                            break
                        else:
                            pass
                                        
                        room_sheet = self.workbook["Nachweisblatt_Raum." + str(j)]

                                        
                        orig_overview_info = []
                        overview_info.append([])

                        overview_info[current_room].append(temp_save_filepath) #Filename
                        orig_overview_info.append(room_sheet['E11']) #Raumname
                        orig_overview_info.append(room_sheet['L90']) #Einhaltung bauliche Grundanforderungen
                        orig_overview_info.append(room_sheet['K98']) #Einhaltung sommerlichen Komfort
                        orig_overview_info.append(room_sheet['B52']) #Benötigte Masnahmen
                        orig_overview_info.append(room_sheet['L51']) #Benötigte g-tot-Werte                 

                        for k in orig_overview_info:
                            if isinstance(k, str):
                                overview_info[current_room].append(str(k).value)
                            elif isinstance(k, float):
                                overview_info[current_room].append(float(k).value)
                            else:
                                overview_info[current_room].append(k.value)
                                        
                        current_room +=1

                except Exception as e:
                    self.status_label.config(text=f"Error: {str(e)}", fg="red")


                save_overview_info_to_excel(overview_info, str(save_folder) + "\DATUM_PHASE_sows_Excel-Ueberisicht.xlsx")

                save_overview_status = tk.Label(root, text="Excel Übersicht erfolgreich gespeichert", fg="green")
                save_overview_status.pack(pady=10)

                print("INFO - Excel Overview saved")
    
    
            save_overview_button = tk.Button(root, text="Excel Übersicht speichern", command=lambda: save_overview(save_folder, save_filepath))  # save Overview Excel
            save_overview_button.pack(pady=10)


            quit_button = tk.Button(root, text="Beenden", command=lambda: root.destroy())  # Quit the Tkinter application
            quit_button.pack(pady=10)

    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

    root = tk.Tk()
    app = Minergie_Excel_Editor(root)
    root.mainloop()
